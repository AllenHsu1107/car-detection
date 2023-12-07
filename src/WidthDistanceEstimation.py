import cv2
import numpy as np
from ultralytics import YOLO
from kalman_filter import kalman_filter
import openpyxl

pixelx = 800
pixely = 600

wb = openpyxl.load_workbook('./testcases/distance.xlsx')

model = YOLO('yolov8n.pt')  # load a pretrained YOLOv8n detection model

objName = 'car'

f = 3040.77957

minwidth = 1.76
maxwidth = 2.06

def DetectObject(objName,img):
    py,px = img.shape[0],img.shape[1]
    output = model(img,verbose=False)
    first_output = output[0]
    names = first_output.names
    width = 0
    for i in first_output.boxes:
        if names[int(i.cls)] == objName:
            x1,y1,x2,y2 = int(i.xyxy[0][0]),int(i.xyxy[0][1]),int(i.xyxy[0][2]),int(i.xyxy[0][3])
            xmid, ymid = (x1+x2)/2,(y1+y2)/2
            if ((ymid/xmid < (py/px*8/3-2/3*py/xmid)) and (ymid/xmid < (-py/px*8/3+2*py/xmid)) and (ymid < 3/4*py)):
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 3)
                width = max(width,x2-x1)
                # print(x2,x1)
    return width

def Zoomin(img, times):
    py,px = img.shape[0],img.shape[1]
    crp_img = img[int(py*(1-1/times)/2):int(py*(1+1/times)/2), int(px*(1-1/times)/2):int(px*(1+1/times)/2)]
    # crp_img = cv2.resize(crp_img,(px,py))
    return crp_img

def EstDistance(pixellength):
    d1 = f*minwidth/pixellength
    d2 = f*maxwidth/pixellength
    return d1,d2

groundtruth_file = open('./testcases/ground_truth.txt')
groundtruth_file.readline()
groundtruth_file.readline()
first = True
frameNumber = 0
s1 = wb['Distance']
for j in range(289):
    groundtruth = float(groundtruth_file.readline())
    path = 'testcases/images/%i.jpg'%j
    frame = cv2.imread(path)
    frameNumber = frameNumber+1
    pixelrange = DetectObject(objName,frame)
    n = 1
    while not pixelrange and n < 4:
        n = n+1
        crp_frame = Zoomin(frame,n)
        pixelrange = DetectObject(objName,crp_frame)
    if pixelrange:
        d1,d2 = EstDistance(pixelrange)
        if first:
            first = False
            x1 = [d1, 0.0]
            p1 = np.diag([1,1])
            x2 = [d2, 0.0]
            p2 = np.diag([1,1])
        x1,p1 = kalman_filter(d1, 0.1*frameNumber, 0.1, 5.0, x1, p1)
        x2,p2 = kalman_filter(d2, 0.1*frameNumber, 0.1, 5.0, x2, p2)
        frameNumber = 0
    else:
        if frameNumber>10:
            first = True
    # frame = cv2.resize(frame,(pixelx,pixely))
    s1.cell(j+1,1).value = d1
    s1.cell(j+1,2).value = x1[0]
    s1.cell(j+1,3).value = d2
    s1.cell(j+1,4).value = x2[0]
    s1.cell(j+1,5).value = groundtruth
    # print(d1,x1[0],d2,x2[0],groundtruth)
# cv2.imshow('ObjectDetection',frame)
# cv2.waitKey(0)
wb.save('./testcases/distance.xlsx')